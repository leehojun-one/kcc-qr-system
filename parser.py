"""
시공발주서(.xlsx) 파서
- KCC 홈씨씨 ERP에서 출력되는 '출고/시공 발주서' 양식을 읽어
  현장정보 + 창호목록을 구조화한다.
- 단가(금액) 컬럼은 의도적으로 제외한다.
- 설치위치가 있는 '실제 창호'만 추출한다 (실측비/철거/양중/부자재 제외).

본사 IT 참고: 양식 좌표가 바뀌면 HEADER_MAP / 창호 컬럼 인덱스만 조정하면 됨.
"""
import json
import sys
from datetime import datetime, date

import openpyxl


def _s(v):
    """셀 값을 안전하게 문자열로."""
    if v is None:
        return ""
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    return str(v).strip()


def parse_order_sheet(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.worksheets[0]

    def c(coord):
        return ws[coord].value

    site = {
        "quote_no": _s(c("D6")),          # 창호견적번호 (마스터 키)
        "order_no": _s(c("H6")),          # 오더번호 (바코드 들어갈 칸)
        "team": _s(c("M6")),              # 소속팀
        "sales_rep": _s(c("M7")),         # 영업자명
        "sales_phone": _s(c("R7")),       # 영업 연락처
        "constructor": _s(c("D8")),       # 시공업체
        "constructor_phone": _s(c("H8")), # 시공업체 연락처
        "vendor": _s(c("M8")),            # 생산업체(가공처)
        "order_date": _s(c("D9")),        # 발주일
        "address": _s(c("L9")),           # 시공주소
        "install_date": _s(c("D10")),     # 시공일
        "windows": [],
    }

    # 창호 블록: 데이터행의 B열(순번)이 숫자(정수 또는 '1' 같은 텍스트). 3행 단위 반복.
    for r in range(14, ws.max_row + 1):
        seq_str = _s(ws.cell(row=r, column=2).value)    # B: 순번
        if not seq_str.isdigit():
            continue
        location = _s(ws.cell(row=r, column=3).value)   # C: 설치위치
        if not location:
            # 설치위치 없음 = 실측비/철거/양중/부자재 → A/S 대상 아님, 제외
            continue
        site["windows"].append({
            "seq": int(seq_str),
            "location": location,
            "model": _s(ws.cell(row=r, column=4).value),   # D: 모델명
            "width": _s(ws.cell(row=r, column=7).value),   # G: 길이(W)
            "height": _s(ws.cell(row=r, column=9).value),  # I: 높이(H)
            "qty": _s(ws.cell(row=r, column=14).value),    # N: 수량
            # 단가(T열)는 읽지 않음
        })

    return site


if __name__ == "__main__":
    path = sys.argv[1] if len(sys.argv) > 1 else "시공발주서.xlsx"
    data = parse_order_sheet(path)
    print(json.dumps(data, ensure_ascii=False, indent=2))
